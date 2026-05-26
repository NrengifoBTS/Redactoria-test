import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import Color
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from io import BytesIO
from typing import Dict, Any, Optional
from pathlib import Path
import logging
import json
from src.api_llm.llm_client import LLMClient
from . import models
from bs4 import BeautifulSoup, NavigableString
import re
from openpyxl.styles.colors import Color


def _rgb_to_hex(rgb_string: str) -> str:
    """
    Convierte formato RGB a hexadecimal ARGB para Excel
    Entrada: 'rgb(59, 130, 246)' o 'rgba(59, 130, 246, 1)'
    Salida: 'FF3b82f6' (con alpha FF para opacidad completa)
    """
    try:
        # Extraer números del RGB
        numbers = re.findall(r'\d+', rgb_string)
        if len(numbers) >= 3:
            r, g, b = int(numbers[0]), int(numbers[1]), int(numbers[2])
            
            # Extraer alpha si existe (rgba)
            alpha = 255  # Opacidad completa por defecto
            if len(numbers) >= 4:
                # Convertir alpha de 0-1 a 0-255
                alpha_float = float(numbers[3]) if '.' in rgb_string else int(numbers[3])
                if alpha_float <= 1:
                    alpha = int(alpha_float * 255)
                else:
                    alpha = int(alpha_float)
            
            # ✅ FORMATO ARGB CORRECTO: Alpha + RGB
            hex_color = f"{alpha:02x}{r:02x}{g:02x}{b:02x}".upper()
            logging.info(f"✅ Convertido RGB({r},{g},{b},alpha={alpha}) -> {hex_color}")
            return hex_color
        else:
            logging.warning(f"No se pudieron extraer 3 números de: {rgb_string}")
            return "FF000000"  # Negro opaco por defecto
    except Exception as e:
        logging.error(f"Error convirtiendo RGB {rgb_string}: {str(e)}")
        return "FF000000"
    
def _normalize_color(color_value: str) -> str:
    """
    Normaliza diferentes formatos de color a hexadecimal ARGB
    Retorna formato ARGB de 8 caracteres para Excel
    """
    if not color_value:
        return "FF000000"
    
    color_value = color_value.strip()
    
    # Si es RGB/RGBA - PRIORIDAD MÁXIMA
    if color_value.startswith('rgb'):
        return _rgb_to_hex(color_value)
    
    # Si ya es hexadecimal con #
    if color_value.startswith('#'):
        hex_value = color_value[1:]
        # Si es RGB (6 chars), agregar alpha FF
        if len(hex_value) == 6:
            return f"FF{hex_value.upper()}"
        # Si ya es ARGB (8 chars)
        elif len(hex_value) == 8:
            return hex_value.upper()
        # Si es formato corto RGB (3 chars), expandir
        elif len(hex_value) == 3:
            expanded = ''.join([c*2 for c in hex_value])
            return f"FF{expanded.upper()}"
    
    # Si es hex sin #
    if len(color_value) == 6 and all(c in '0123456789ABCDEFabcdef' for c in color_value):
        return f"FF{color_value.upper()}"
    
    if len(color_value) == 8 and all(c in '0123456789ABCDEFabcdef' for c in color_value):
        return color_value.upper()
    
    # Colores CSS nombrados
    css_colors = {
        'red': 'FFFF0000',
        'blue': 'FF0000FF',
        'green': 'FF008000',
        'black': 'FF000000',
        'white': 'FFFFFFFF',
        'yellow': 'FFFFFF00',
        'orange': 'FFFFA500',
        'purple': 'FF800080',
        'pink': 'FFFFC0CB',
        'gray': 'FF808080',
        'grey': 'FF808080',
    }
    
    if color_value.lower() in css_colors:
        return css_colors[color_value.lower()]
    
    logging.warning(f"Formato de color no reconocido: {color_value}, usando negro")
    return "FF000000"

def _parse_html_to_rich_text(html_content: str):
    """
    Versión corregida con mejor manejo de colores
    """
    from bs4 import BeautifulSoup, NavigableString
    from openpyxl.cell.rich_text import TextBlock, CellRichText
    from openpyxl.cell.text import InlineFont
    
    logging.info(f"=== PARSING HTML TO RICH TEXT (FIXED) ===")
    logging.info(f"Input: {html_content}")
    
    if not html_content or html_content.strip() == '':
        return ""
    
    if '<' not in html_content:
        return html_content
    
    try:
        invalid_tags = ['g', 'mo', 'alquila', 'renta', 'auto', 'viaje']
        for tag in invalid_tags:
            html_content = re.sub(rf'<{tag}[^>]*>', '<span>', html_content, flags=re.IGNORECASE)
            html_content = re.sub(rf'</{tag}>', '</span>', html_content, flags=re.IGNORECASE)
        html_content = re.sub(r'\s+\w+=""', '', html_content)  # Limpiar attrs vacíos
        
        soup = BeautifulSoup(html_content, 'html.parser')
        text_blocks = []
        
        def process_element(element, current_styles=None):
            if current_styles is None:
                current_styles = {}
            
            if isinstance(element, NavigableString):
                text = str(element)
                if text:
                    text_blocks.append({
                        'text': text,
                        'styles': current_styles.copy()
                    })
                return
            
            element_styles = current_styles.copy()
            
            # Tags de formato
            if element.name in ['strong', 'b']:
                element_styles['bold'] = True
            if element.name in ['em', 'i']:
                element_styles['italic'] = True
            if element.name == 'u':
                element_styles['underline'] = True
            
            # Analizar style attribute
            if element.get('style'):
                style_attr = element.get('style')
                
                # ✅ MEJORADO: Buscar color con más precisión
                # Primero buscar rgba
                rgba_match = re.search(r'color:\s*rgba?\s*\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*(?:,\s*[\d.]+)?\s*\)', style_attr)
                if rgba_match:
                    r, g, b = rgba_match.groups()
                    rgb_string = f"rgb({r}, {g}, {b})"
                    element_styles['color'] = _normalize_color(rgb_string)
                    logging.info(f"✅ Color RGB extraído: {rgb_string} -> {element_styles['color']}")
                else:
                    # Buscar color hex o nombrado
                    color_match = re.search(r'color:\s*([#\w]+)', style_attr)
                    if color_match:
                        color_value = color_match.group(1).strip()
                        element_styles['color'] = _normalize_color(color_value)
                        logging.info(f"✅ Color extraído: {color_value} -> {element_styles['color']}")
                
                # Bold/italic en style
                if re.search(r'font-weight:\s*bold', style_attr):
                    element_styles['bold'] = True
                if re.search(r'font-style:\s*italic', style_attr):
                    element_styles['italic'] = True
            
            # Procesar hijos
            for child in element.children:
                process_element(child, element_styles)
        
        # Extraer bloques
        for child in soup.children:
            process_element(child, {})
        
        logging.info(f"Bloques extraídos: {len(text_blocks)}")
        for i, block in enumerate(text_blocks[:5]):  # Mostrar primeros 5
            logging.info(f"  Bloque {i}: text='{block['text'][:30]}...' styles={block['styles']}")
        
        # Consolidar bloques con mismo estilo
        consolidated_blocks = []
        i = 0
        while i < len(text_blocks):
            current_block = text_blocks[i]
            current_text = current_block['text']
            current_styles = current_block['styles']
            
            j = i + 1
            while j < len(text_blocks):
                next_block = text_blocks[j]
                if next_block['styles'] == current_styles:
                    current_text += next_block['text']
                    j += 1
                else:
                    break
            
            if current_text:
                consolidated_blocks.append({
                    'text': current_text,
                    'styles': current_styles
                })
            
            i = j
        
        logging.info(f"Bloques consolidados: {len(consolidated_blocks)}")
        
        # Si solo hay un bloque sin estilos, retornar texto simple
        if len(consolidated_blocks) == 1 and not consolidated_blocks[0]['styles']:
            return consolidated_blocks[0]['text']
        
        # Crear RichText
        if len(consolidated_blocks) > 1 or (len(consolidated_blocks) == 1 and consolidated_blocks[0]['styles']):
            rich_text_parts = []
            
            for block in consolidated_blocks:
                text = block['text']
                styles = block['styles']
                
                font_kwargs = {
                    'rFont': 'Calibri',
                    'sz': 11,
                    'b': styles.get('bold', False),
                    'i': styles.get('italic', False),
                    'u': 'single' if styles.get('underline', False) else None
                }
                
                if 'color' in styles:
                    try:
                        color_argb = styles['color']
                        # Verificar que sea formato ARGB válido
                        if len(color_argb) == 8:
                            font_kwargs['color'] = Color(rgb=color_argb)
                            logging.info(f"✅ Color aplicado a TextBlock: {color_argb}")
                        else:
                            logging.warning(f"⚠️ Color inválido (no es ARGB): {color_argb}")
                    except Exception as e:
                        logging.error(f"❌ Error aplicando color {styles['color']}: {str(e)}")
                
                inline_font = InlineFont(**font_kwargs)
                text_block = TextBlock(inline_font, text)
                rich_text_parts.append(text_block)
                
                logging.info(f"✅ TextBlock creado: '{text[:30]}...' con estilos {styles}")
            
            rich_text = CellRichText(rich_text_parts)
            logging.info(f"✅ RichText final con {len(rich_text_parts)} bloques")
            return rich_text
        
        # Fallback
        if consolidated_blocks:
            return ''.join([block['text'] for block in consolidated_blocks])
        else:
            return ""
        
    except Exception as e:
        logging.error(f"❌ Error parsing HTML to RichText: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return re.sub(r'<[^>]+>', '', html_content)

def debug_cell_data_structure(export_request: models.ExportExcelRequest):
    """Debug para ver la estructura exacta de los datos"""
    logging.info("=== DEBUG: ESTRUCTURA DE DATOS ===")
    
    # Debug cell_data
    if export_request.cell_data:
        logging.info(f"cell_data existe, {len(export_request.cell_data)} celdas")
        for i, (key, value) in enumerate(export_request.cell_data.items()):
            if i < 5:  # Solo mostrar las primeras 5
                logging.info(f"cell_data[{key}] = {value}")
                logging.info(f"  - type: {type(value)}")
                logging.info(f"  - value: {getattr(value, 'value', 'NO_VALUE_ATTR')}")
    else:
        logging.info("cell_data es None")
    
    # Debug template_data
    template_data = export_request.template_config.templateData
    if template_data:
        logging.info(f"templateData existe, {len(template_data)} celdas")
        for i, (key, value) in enumerate(template_data.items()):
            if i < 5:  # Solo mostrar las primeras 5
                logging.info(f"templateData[{key}] = {value}")
                logging.info(f"  - type: {type(value)}")
                logging.info(f"  - value: {getattr(value, 'value', 'NO_VALUE_ATTR')}")
    else:
        logging.info("templateData es None")

def _parse_html_to_excel_format(html_content: str) -> dict:
    """
    Parsea HTML con formato y extrae texto plano + estilos para Excel
    """
    logging.info(f"=== PARSING HTML ===")
    logging.info(f"Input: {html_content[:200]}...")  # Solo los primeros 200 caracteres
    
    if not html_content or html_content.strip() == '':
        logging.info("HTML vacío, retornando texto vacío")
        return {"text": "", "styles": {}}
    
    try:
        # Si no hay tags HTML, retornar como texto plano
        if '<' not in html_content:
            logging.info("No hay tags HTML, retornando texto plano")
            return {"text": html_content, "styles": {}}
        
        # Parsear HTML
        soup = BeautifulSoup(html_content, 'html.parser')
        
        # Extraer texto plano
        plain_text = soup.get_text()
        logging.info(f"Texto extraído: {plain_text}")
        
        # Analizar estilos
        styles = {}
        
        # 1. Negritas
        strong_elements = soup.find_all(['strong', 'b'])
        if strong_elements or 'font-weight:bold' in html_content or 'font-weight: bold' in html_content:
            styles['bold'] = True
            logging.info("✅ NEGRITA detectada")
        
        # 2. Colores
        color_spans = soup.find_all('span', style=True)
        for span in color_spans:
            style_attr = span.get('style', '')
            color_match = re.search(r'color:\s*([^;"\s]+)', style_attr)
            if color_match:
                color_value = color_match.group(1).strip()
                if color_value.startswith('#'):
                    color_value = color_value[1:]
                styles['color'] = color_value
                logging.info(f"✅ COLOR detectado: {color_value}")
                break
        
        # También buscar en el HTML completo
        if 'color' not in styles:
            color_match = re.search(r'color:\s*([^;"\s]+)', html_content)
            if color_match:
                color_value = color_match.group(1).strip()
                if color_value.startswith('#'):
                    color_value = color_value[1:]
                styles['color'] = color_value
                logging.info(f"✅ COLOR detectado (regex): {color_value}")
        
        # 3. Cursiva
        if soup.find(['em', 'i']) or 'font-style:italic' in html_content:
            styles['italic'] = True
            logging.info("✅ CURSIVA detectada")
        
        logging.info(f"Estilos finales: {styles}")
        
        result = {
            "text": plain_text,
            "styles": styles
        }
        
        logging.info(f"Resultado del parsing: {result}")
        return result
        
    except Exception as e:
        logging.error(f"Error parsing HTML content: {str(e)}")
        # Fallback: intentar extraer texto básico
        clean_text = re.sub(r'<[^>]+>', '', html_content)
        return {"text": clean_text, "styles": {}}
    
# Funciones de template integradas directamente
def get_template_path(proyecto: str, categoria: str) -> Optional[Path]:
    """
    Busca el template Excel correspondiente al proyecto y categoría
    """
    # Buscar en varias ubicaciones posibles
    possible_paths = [
        Path(__file__).parent.parent / "extemp" / "excel",
        Path(__file__).parent.parent / "templates" / "excel", 
        Path(__file__).parent / "templates",
        Path("templates/excel"),
        Path("extemp/excel")
    ]
    
    patterns = [
        f"{proyecto}_{categoria}.xlsx",
        f"{proyecto}_default.xlsx", 
        f"{categoria}.xlsx",
        "default.xlsx"
    ]
    
    for base_path in possible_paths:
        if base_path.exists():
            for pattern in patterns:
                template_path = base_path / pattern
                if template_path.exists():
                    logging.info(f"Template encontrado: {template_path}")
                    return template_path
    
    logging.warning(f"No se encontró template para {proyecto}_{categoria}")
    return None

def get_template_info(proyecto: str, categoria: str) -> dict:
    """Obtiene información del template que se usará"""
    template_path = get_template_path(proyecto, categoria)
    
    if template_path:
        return {
            "found": True,
            "path": str(template_path),
            "filename": template_path.name,
            "size": template_path.stat().st_size,
            "proyecto": proyecto,
            "categoria": categoria
        }
    else:
        return {
            "found": False,
            "path": None,
            "filename": None,
            "proyecto": proyecto,
            "categoria": categoria,
            "message": f"No se encontró template para {proyecto}_{categoria}"
        }

def create_excel_from_template(export_request: models.ExportExcelRequest) -> BytesIO:
    """
    Crea un archivo Excel basado en template existente o crea uno nuevo
    """
    try:
        # Buscar template existente
        template_path = get_template_path(
            export_request.template_info.proyecto,
            export_request.template_info.categoria
        )
        
        if template_path and template_path.exists():
            # Cargar template existente
            wb = openpyxl.load_workbook(template_path)
            
            logging.info(f"Hojas disponibles en el template: {wb.sheetnames}")
            logging.info(f"Template cargado desde: {template_path}")
            
            # Trabajar con la hoja "traducciones" (común para todos)
            if "traducciones" in wb.sheetnames:
                ws_traducciones = wb["traducciones"]
                _update_template_with_data(ws_traducciones, export_request)
                ws_traducciones.title = f"{export_request.template_info.name}"
                logging.info("Hoja 'traducciones' procesada")
            else:
                # Si no existe la hoja traducciones, usar la activa
                ws = wb.active
                _update_template_with_data(ws, export_request)
                ws.title = f"{export_request.template_info.name}"
                logging.info("Usando hoja activa como traducciones")
            
            # Procesar hoja "imagenes" según el template específico
            _process_imagenes_by_template(wb, export_request)
            
        else:
            # No existe template
            logging.info("No se encontró template")
            raise Exception("Template no encontrado")
        
        # Guardar en memoria
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logging.info(f"Excel created successfully for template: {export_request.template_info.name}")
        return excel_buffer
        
    except Exception as e:
        logging.error(f"Error creating Excel file: {str(e)}")
        raise

def _process_imagenes_by_template(wb, export_request: models.ExportExcelRequest):
    """
    Procesa la hoja imagenes según el template específico
    """
    template_name = export_request.template_info.name
    
    # Verificar si existe hoja imagenes
    if "imagenes" not in wb.sheetnames:
        logging.info(f"Template '{template_name}' no tiene hoja 'imagenes'")
        return
      
    ws_imagenes = wb["imagenes"]
    logging.info(f"Procesando hoja 'imagenes' para template: {template_name}")
    
    # Router: ejecutar lógica específica según el template
    if template_name == "Template Autos":
        _populate_imagenes_viajemos_autos(ws_imagenes, export_request)
    elif template_name == "Template Ciudad":
        _populate_imagenes_viajemos_ciudad(ws_imagenes, export_request)    
    elif template_name == "otro_template":
        # Para otros templates, usar la lógica antigua de un solo bloque
        keywords_data = _generate_keywords_with_llm(export_request)
        if keywords_data and "keywords" in keywords_data:
            keywords = keywords_data["keywords"]
            _fill_keywords_otro_template(ws_imagenes, keywords)
    else:
        logging.info(f"Template '{template_name}' no tiene lógica específica para hoja imagenes")
        
# =================== GENERACIÓN DE KEYWORDS (COMÚN) ===================

def _build_llm_prompt_with_title(titulo: str) -> str:
    """
    Construye prompt con título específico
    """
    prompt = f"""Eres un experto en SEO que genera keywords. Basándote en el título proporcionado, crea keywords siguiendo estos parámetros:
Tono: Jovial, fresco
Verbos principales: alquiler, renta, arriendo (conjugalos según el contexto, usar plural y singular según corresponda)
Público objetivo: USA y LATAM
Generar solo 2 keywords en español basadas en el título, mezclando: términos del título, variaciones de verbos, ubicaciones geográficas si es necesario, o frases comerciales/long-tail
Usar conectores naturales (de, en, para, con, cómo, dónde, etc.)
y traducir las keywords al inglés y portugués, manteniendo el mismo formato.
FORMATO DE RESPUESTA: Responde únicamente con este JSON (sin explicaciones adicionales):
{{
  "keywords": {{
    "español": [
      "keyword 1 en español",
      "keyword 2 en español"
    ],
    "inglés": [
      "keyword 1 in English",
      "keyword 2 in English"
    ],
    "portugués": [
      "keyword 1 en portugues",
      "keyword 2 en portugues"
    ]
  }}
}}

TÍTULO A PROCESAR: {titulo}"""
    
    return prompt

def _clean_llm_response(response: str) -> str:
    """Limpia la respuesta del LLM para extraer solo el JSON"""
    try:
        # Si la respuesta viene envuelta en markdown, extraer el JSON
        if "```json" in response:
            # Encontrar el inicio y fin del bloque JSON
            start = response.find("```json") + 7
            end = response.find("```", start)
            if end != -1:
                json_content = response[start:end].strip()
                return json_content
        
        # Si no tiene markdown, devolver la respuesta tal como está
        return response.strip()
        
    except Exception as e:
        logging.error(f"Error limpiando respuesta del LLM: {str(e)}")
        return response

def _extract_title_from_block(export_request: models.ExportExcelRequest, block_id: str) -> Optional[str]:
    """Extrae el título de un bloque específico usando blocks_metadata y cell_data"""
    try:
        logging.info(f"Intentando extraer título del bloque {block_id}")
        
        # Verificar diferentes formas de acceder a los datos
        blocks_metadata = None
        cell_data = None
        
        # Intentar diferentes formas de acceder a blocks_metadata
        if hasattr(export_request, 'blocks_metadata') and export_request.blocks_metadata:
            blocks_metadata = export_request.blocks_metadata
        elif hasattr(export_request, 'template_config') and hasattr(export_request.template_config, 'blocks_metadata'):
            blocks_metadata = export_request.template_config.blocks_metadata
        
        # Intentar diferentes formas de acceder a cell_data
        if hasattr(export_request, 'cell_data') and export_request.cell_data:
            cell_data = export_request.cell_data
        elif hasattr(export_request, 'template_config') and hasattr(export_request.template_config, 'cell_data'):
            cell_data = export_request.template_config.cell_data
        
        if not blocks_metadata:
            logging.warning("blocks_metadata no encontrado en ninguna ubicación")
            return None
            
        if not cell_data:
            logging.warning("cell_data no encontrado en ninguna ubicación")
            return None
        
        logging.info(f"blocks_metadata encontrado: {blocks_metadata}")
        logging.info(f"cell_data claves disponibles: {list(cell_data.keys())[:10]}")
        
        if block_id not in blocks_metadata:
            logging.warning(f"Bloque {block_id} no encontrado. Bloques disponibles: {list(blocks_metadata.keys())}")
            return None
        
        block_info = blocks_metadata[block_id]
        logging.info(f"Info del bloque {block_id}: {block_info}")
        
        # CORRECCIÓN: Acceder como atributo del objeto, no como diccionario
        if not hasattr(block_info, 'titleRow'):
            logging.warning(f"titleRow no encontrado en bloque {block_id}")
            return None
            
        title_row = block_info.titleRow  # Cambio: usar punto en lugar de corchetes
        title_cell_key = f"{title_row}-3"
        logging.info(f"Buscando título en celda: {title_cell_key}")
        
        if title_cell_key not in cell_data:
            logging.warning(f"Celda {title_cell_key} no encontrada en cell_data")
            # DEBUG: Mostrar claves similares
            similar_keys = [k for k in cell_data.keys() if k.startswith(f"{title_row}-")]
            logging.info(f"Claves similares encontradas: {similar_keys}")
            return None
            
        title_cell = cell_data[title_cell_key]
        logging.info(f"Datos de la celda {title_cell_key}: {title_cell}")
        
        # CORRECCIÓN: También verificar si title_cell es un objeto
        if hasattr(title_cell, 'value'):
            titulo = title_cell.value  # Es un objeto
        elif isinstance(title_cell, dict) and 'value' in title_cell:
            titulo = title_cell['value']  # Es un diccionario
        else:
            logging.warning(f"Estructura de celda no reconocida: {title_cell}")
            return None
        
        if titulo:
            logging.info(f"Título extraído del bloque {block_id}: {titulo}")
            return str(titulo)
        else:
            logging.warning(f"Título vacío en celda {title_cell_key}")
            return None
            
    except Exception as e:
        logging.error(f"Error extrayendo título del bloque {block_id}: {str(e)}")
        return None

def _fill_single_keyword_block(ws_imagenes, keywords: Dict, start_row: int, block_name: str):
    """
    Llena un solo bloque de keywords empezando en start_row
    """
    try:
        logging.info(f"Llenando {block_name} con keywords: {type(keywords)}")
        
        row = start_row
        
        # Verificar que keywords es un diccionario
        if not isinstance(keywords, dict):
            logging.error(f"Keywords para {block_name} no es un diccionario: {keywords}")
            return
        
        spanish_keywords = keywords.get("español", [])
        english_keywords = keywords.get("inglés", [])
        portuguese_keywords = keywords.get("portugués", [])
        
        # Determinar cuántas keywords por idioma tenemos
        max_keywords = min(len(spanish_keywords), len(english_keywords), len(portuguese_keywords), 2)
        
        # Llenar en orden intercalado: español, inglés, portugués para cada posición
        for i in range(max_keywords):
            if i < len(spanish_keywords):
                ws_imagenes[f"D{row}"].value = spanish_keywords[i]
                row += 1
            
            if i < len(english_keywords):
                ws_imagenes[f"D{row}"].value = english_keywords[i]
                row += 1
            
            if i < len(portuguese_keywords):
                ws_imagenes[f"D{row}"].value = portuguese_keywords[i]
                row += 1
        
        logging.info(f"Keywords llenadas para {block_name} (filas D{start_row}-D{row-1})")
        
    except Exception as e:
        logging.error(f"Error llenando bloque {block_name}: {str(e)}")
               
def _generate_keywords_for_block(export_request: models.ExportExcelRequest, block_id: str) -> Optional[Dict]:
    """
    Genera keywords para un bloque específico
    """
    try:
        titulo = _extract_title_from_block(export_request, block_id)
        
        if not titulo:
            titulo = export_request.template_info.name
            logging.warning(f"No se encontró título en bloque {block_id}, usando template name como fallback")
        
        prompt = _build_llm_prompt_with_title(titulo)
        
        llm_client = LLMClient()
        response = llm_client.generate(
            prompt=prompt,
            system_message="Eres un experto en generación de keywords para contenido multiidioma. Siempre responde con JSON válido."
        )
        
        if response:
            cleaned_response = _clean_llm_response(response)
            keywords_data = json.loads(cleaned_response)
            logging.info(f"Keywords data completa para bloque {block_id}: {keywords_data}")
            
            if "keywords" in keywords_data:
                keywords = keywords_data["keywords"]
                logging.info(f"Keywords extraídas para bloque {block_id}: {keywords}")
                return keywords  
            else:
                logging.error(f"No se encontró 'keywords' en la respuesta del bloque {block_id}")
                return None
        
        return None
        
    except json.JSONDecodeError as e:
        logging.error(f"Error parseando JSON para bloque {block_id}: {str(e)}")
        logging.error(f"Respuesta limpia: {cleaned_response if 'cleaned_response' in locals() else 'N/A'}")
        return None
    except Exception as e:
        logging.error(f"Error generando keywords para bloque {block_id}: {str(e)}")
        return None
    
def _extract_title_from_cell_reference(export_request: models.ExportExcelRequest, cell_reference: str) -> Optional[str]:
    """
    Extrae título de una referencia de celda específica (ej: "30-3")
    """
    try:
        # Extraer cell_data
        cell_data = None
        if hasattr(export_request, 'cell_data'):
            cell_data = export_request.cell_data
        elif hasattr(export_request, 'template_config') and hasattr(export_request.template_config, 'cell_data'):
            cell_data = export_request.template_config.cell_data
        
        if not cell_data:
            logging.warning("cell_data no encontrado")
            return None
        
        logging.info(f"Buscando título en celda: {cell_reference}")
        
        if cell_reference not in cell_data:
            logging.warning(f"Celda {cell_reference} no encontrada en cell_data")
            return None
        
        title_cell = cell_data[cell_reference]
        
        # Extraer valor dependiendo de si es objeto o diccionario
        if hasattr(title_cell, 'value'):
            titulo = title_cell.value
        elif isinstance(title_cell, dict) and 'value' in title_cell:
            titulo = title_cell['value']
        else:
            logging.warning(f"Estructura de celda no reconocida: {title_cell}")
            return None
        
        if titulo:
            logging.info(f"Título extraído de {cell_reference}: {titulo}")
            return str(titulo)
        else:
            logging.warning(f"Título vacío en celda {cell_reference}")
            return None
        
    except Exception as e:
        logging.error(f"Error extrayendo título de {cell_reference}: {str(e)}")
        return None
    
def _generate_keywords_with_title(titulo: str, block_name: str) -> Optional[Dict]:
    """
    Genera keywords con un título específico
    """
    try:
        logging.info(f"Generando keywords para {block_name} con título: {titulo}")
        
        prompt = _build_llm_prompt_with_title(titulo)
        
        llm_client = LLMClient()
        response = llm_client.generate(
            prompt=prompt,
            system_message="Eres un experto en generación de keywords para contenido multiidioma. Siempre responde con JSON válido."
        )
        
        if response:
            cleaned_response = _clean_llm_response(response)
            keywords_data = json.loads(cleaned_response)
            
            if "keywords" in keywords_data:
                keywords = keywords_data["keywords"]
                logging.info(f"Keywords generadas para {block_name}")
                return keywords
            else:
                logging.error(f"No se encontró 'keywords' en la respuesta de {block_name}")
                return None
        
        return None
        
    except Exception as e:
        logging.error(f"Error generando keywords para {block_name}: {str(e)}")
        return None

# =================== LLENADO ESPECÍFICO POR TEMPLATE ===================
 
def _fill_keywords_otro_template(ws_imagenes, keywords: Dict):
    """
    Llena celdas para otro_template (ejemplo: diferente formato)
    """
    try:
        # TODO: Implementar lógica específica
        # Ejemplo: podría llenar en columnas diferentes, formato diferente, etc.
        
        # Ejemplo: llenar en formato agrupado por idioma
        ws_imagenes["E2"].value = "Español:"
        ws_imagenes["E3"].value = ", ".join(keywords.get("español", []))
        ws_imagenes["E4"].value = "Inglés:"
        ws_imagenes["E5"].value = ", ".join(keywords.get("inglés", []))
        ws_imagenes["E6"].value = "Portugués:"
        ws_imagenes["E7"].value = ", ".join(keywords.get("portugués", []))
        
        logging.info("Keywords otro_template agregadas")
        
    except Exception as e:
        logging.error(f"Error llenando keywords otro_template: {str(e)}")




def _populate_imagenes_viajemos_autos(ws_imagenes, export_request: models.ExportExcelRequest):
    """
    Lógica específica para template viajemos_autos - múltiples bloques
    """
    try:
        logging.info("Procesando hoja imagenes para viajemos_autos...")
        
        keywords_blocks = []
        block_positions = []
        
        # Bloque 1: extraer título del bloque "1"
        logging.info("Generando keywords para Bloque 1...")
        keywords_block1 = _generate_keywords_for_block(export_request, "1")
        if keywords_block1:
            keywords_blocks.append(keywords_block1)
            block_positions.append({"start_row": 2, "name": "Bloque 1"})
        
        # Bloque 3: extraer título del bloque "3" y generar 2 peticiones
        logging.info("Generando keywords para Bloque 3 (primera petición)...")
        keywords_block3a = _generate_keywords_for_block(export_request, "3")
        if keywords_block3a:
            keywords_blocks.append(keywords_block3a)
            block_positions.append({"start_row": 10, "name": "Bloque 3a"})
        
        logging.info("Generando keywords para Bloque 3 (segunda petición)...")
        keywords_block3b = _generate_keywords_for_block(export_request, "3")
        if keywords_block3b:
            keywords_blocks.append(keywords_block3b)
            block_positions.append({"start_row": 18, "name": "Bloque 3b"})
        
        # Bloque 5: dinámico basado en contentMapping
        logging.info("Procesando Bloque 5 dinámico...")
        block5_data = _process_dynamic_block5(export_request)
        
        for i, (keywords, position) in enumerate(block5_data):
            if keywords:
                keywords_blocks.append(keywords)
                block_positions.append(position)
                
        # Bloque 6: dinámico basado en contentMapping
        logging.info("Procesando Bloque 6 dinámico...")
        block6_data = _process_dynamic_block6(export_request)

        for i, (keywords, position) in enumerate(block6_data):
            if keywords:
                keywords_blocks.append(keywords)
                block_positions.append(position)
        
        # Llenar todos los bloques
        if keywords_blocks:
            _fill_keywords_viajemos_autos_dynamic(ws_imagenes, keywords_blocks, block_positions)
            logging.info("Hoja 'imagenes' poblada para viajemos_autos con múltiples bloques")
        else:
            logging.warning("No se pudieron generar keywords para ningún bloque")
            
    except Exception as e:
        logging.error(f"Error poblando hoja imagenes viajemos_autos: {str(e)}")

def _populate_imagenes_viajemos_ciudad(ws_imagenes, export_request: models.ExportExcelRequest):
    """
    Lógica específica para template viajemos_autos - múltiples bloques
    """
    try:
        logging.info("Procesando hoja imagenes para viajemos_ciudad...")
        
        keywords_blocks = []
        block_positions = []
        
        # Bloque 1: extraer título del bloque "1"
        logging.info("Generando keywords para Bloque 1...")
        keywords_block1 = _generate_keywords_for_block(export_request, "1")
        if keywords_block1:
            keywords_blocks.append(keywords_block1)
            block_positions.append({"start_row": 2, "name": "Bloque 1"})
        
        # Bloque 3: extraer título del bloque "3" y generar 2 peticiones
        logging.info("Generando keywords para Bloque 3 (primera petición)...")
        keywords_block3a = _generate_keywords_for_block(export_request, "3")
        if keywords_block3a:
            keywords_blocks.append(keywords_block3a)
            block_positions.append({"start_row": 10, "name": "Bloque 3a"})
        
        
        # Bloque 5: dinámico basado en contentMapping
        logging.info("Procesando Bloque 5 dinámico...")
        block5_data = _process_dynamic_block5(export_request)
        
        for i, (keywords, position) in enumerate(block5_data):
            if keywords:
                keywords_blocks.append(keywords)
                block_positions.append(position)
                
        # Bloque 6: dinámico basado en contentMapping
        logging.info("Procesando Bloque 6 dinámico...")
        block6_data = _process_dynamic_block6(export_request)

        for i, (keywords, position) in enumerate(block6_data):
            if keywords:
                keywords_blocks.append(keywords)
                block_positions.append(position)
        
        # Llenar todos los bloques
        if keywords_blocks:
            _fill_keywords_viajemos_autos_dynamic(ws_imagenes, keywords_blocks, block_positions)
            logging.info("Hoja 'imagenes' poblada para viajemos_autos con múltiples bloques")
        else:
            logging.warning("No se pudieron generar keywords para ningún bloque")
            
    except Exception as e:
        logging.error(f"Error poblando hoja imagenes viajemos_autos: {str(e)}")


def _process_dynamic_block5(export_request: models.ExportExcelRequest) -> list:
    """
    Procesa el bloque 5 dinámico basado en contentMapping
    """
    try:
        logging.info("=== PROCESANDO BLOQUE 5 DINÁMICO ===")
        
        template_data = export_request.template_config.model_dump()
        blocks_data = template_data['blocks_metadata']
        
        if '5' not in blocks_data:
            logging.error("Bloque 5 no encontrado")
            return []
        
        def buscar_contentmapping_bloque5(data, ruta=""):
            if isinstance(data, dict):
                for k, v in data.items():
                    if k == '5' and isinstance(v, dict) and 'contentMapping' in v:
                        return v['contentMapping']
                    elif isinstance(v, dict):
                        result = buscar_contentmapping_bloque5(v, f"{ruta}.{k}")
                        if result:
                            return result
            return None
        
        content_mapping = buscar_contentmapping_bloque5(template_data)
        
        if not content_mapping:
            logging.error("contentMapping del bloque 5 no encontrado")
            return []
        
        # Extraer todos los desc_# del contentMapping (excluyendo "desc" sin número)
        desc_items = []
        for key, value in content_mapping.items():
            if key.startswith('desc_'):  # Solo desc_1, desc_2, etc.
                desc_number = int(key.split('_')[1])
                desc_items.append((desc_number, key, value))
        
        # Ordenar por número de desc
        desc_items.sort(key=lambda x: x[0])
        logging.info(f"Encontrados {len(desc_items)} desc items para bloque 5: {[item[1] for item in desc_items]}")
        
        # Procesar cada desc_#
        results = []
        for desc_number, desc_key, cell_reference in desc_items:
            logging.info(f"Procesando {desc_key} con referencia {cell_reference}")
            
            # Extraer título específico de esta desc
            titulo = _extract_title_from_cell_reference(export_request, cell_reference)
            
            if titulo:
                # Generar keywords con este título específico
                keywords = _generate_keywords_with_title(titulo, desc_key)
                start_row = 26 + (desc_number - 1) * 8
                position = {"start_row": start_row, "name": f"Bloque 5 {desc_key}"}
                
                results.append((keywords, position))
            else:
                logging.warning(f"No se pudo extraer título para {desc_key}")
        
        logging.info(f"=== FIN PROCESANDO BLOQUE 5 - {len(results)} items procesados ===")
        return results
        
    except Exception as e:
        logging.error(f"Error procesando bloque 5 dinámico: {str(e)}")
        return []
   
def _process_dynamic_block6(export_request: models.ExportExcelRequest) -> list:
    """
    Procesa el bloque 6 dinámico basado en contentMapping
    """
    try:
        logging.info("=== PROCESANDO BLOQUE 6 DINÁMICO ===")
        
        template_data = export_request.template_config.model_dump()
        blocks_data = template_data['blocks_metadata']
        
        if '6' not in blocks_data:
            logging.error("Bloque 6 no encontrado")
            return []
        
        # Acceder de forma diferente - debe estar en la estructura anidada
        # Buscar recursivamente en template_data
        def buscar_contentmapping_bloque6(data, ruta=""):
            if isinstance(data, dict):
                for k, v in data.items():
                    if k == '6' and isinstance(v, dict) and 'contentMapping' in v:
                        return v['contentMapping']
                    elif isinstance(v, dict):
                        result = buscar_contentmapping_bloque6(v, f"{ruta}.{k}")
                        if result:
                            return result
            return None
        
        content_mapping = buscar_contentmapping_bloque6(template_data)
        
        if not content_mapping:
            logging.error("contentMapping del bloque 6 no encontrado")
            return []
        
        logging.info(f"contentMapping del bloque 6 encontrado: {content_mapping}")
        
        # Extraer todos los desc_# del contentMapping (excluyendo "desc" sin número)
        desc_items = []
        for key, value in content_mapping.items():
            if key.startswith('desc_'):  # Solo desc_1, desc_2, etc.
                desc_number = int(key.split('_')[1])
                desc_items.append((desc_number, key, value))
        
        # Ordenar por número de desc
        desc_items.sort(key=lambda x: x[0])
        logging.info(f"Encontrados {len(desc_items)} desc items para bloque 6: {[item[1] for item in desc_items]}")
        
        # Procesar cada desc_#
        results = []
        for desc_number, desc_key, cell_reference in desc_items:
            logging.info(f"Procesando {desc_key} con referencia {cell_reference}")
            
            # Extraer título específico de esta desc
            titulo = _extract_title_from_cell_reference(export_request, cell_reference)
            
            if titulo:
                # Generar keywords con este título específico
                keywords = _generate_keywords_with_title(titulo, desc_key)
                
                base_start_row = 170  
                start_row = base_start_row + (desc_number - 1) * 8
                position = {"start_row": start_row, "name": f"Bloque 6 {desc_key}"}
                
                results.append((keywords, position))
            else:
                logging.warning(f"No se pudo extraer título para {desc_key}")
        
        logging.info(f"=== FIN PROCESANDO BLOQUE 6 - {len(results)} items procesados ===")
        return results
        
    except Exception as e:
        logging.error(f"Error procesando bloque 6 dinámico: {str(e)}")
        return []

def _fill_keywords_viajemos_autos_dynamic(ws_imagenes, keywords_blocks: list, block_positions: list):
    """
    Llena múltiples bloques de keywords usando posiciones dinámicas
    """
    try:
        # Llenar cada bloque con sus keywords correspondientes
        for keywords, position in zip(keywords_blocks, block_positions):
            _fill_single_keyword_block(ws_imagenes, keywords, position["start_row"], position["name"])
        
        logging.info(f"Llenados {len(keywords_blocks)} bloques de keywords dinámicos")
        
    except Exception as e:
        logging.error(f"Error llenando múltiples bloques dinámicos: {str(e)}")
        
# =================== FUNCIONES ORIGINALES  ===================

def _find_merge_master(ws, row, col): 
    """Encuentra la celda principal de un rango combinado"""
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and 
            merged_range.min_col <= col <= merged_range.max_col):
            # Retornar la celda principal (esquina superior izquierda)
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return None

def _update_template_with_data(ws, export_request: models.ExportExcelRequest):
    """Actualiza un template existente con los datos nuevos - SOLO COLUMNAS D, E, F"""
    try:
        logging.info("=== INICIANDO ACTUALIZACIÓN DE TEMPLATE ===")
        
        # AGREGAR DEBUG AQUÍ
        debug_cell_data_structure(export_request)
        
        # Solo actualizar las celdas con datos, manteniendo el formato del template
        if export_request.cell_data:
            logging.info("Actualizando con cell_data...")
            _update_cell_data_safe(ws, export_request.cell_data)
        
        # Actualizar datos del template base
        logging.info("Actualizando con templateData...")
        _populate_template_data_safe(ws, export_request.template_config.templateData)
        
        logging.info("✅ Template existente actualizado con datos (solo columnas D, E, F)")
        
    except Exception as e:
        logging.error(f"❌ Error actualizando template existente: {str(e)}")
        import traceback
        logging.error(f"Traceback: {traceback.format_exc()}")
            
def _populate_template_data_safe(ws, template_data: Dict[str, models.CellData]):
    """Pobla las celdas con los datos del template - CON RICH TEXT"""
    for cell_key, cell_data in template_data.items():
        try:
            # Parsear coordenadas (formato: "row-col")
            row, col = map(int, cell_key.split('-'))
            
            # Ajustar para que las filas empiecen en 2 (después del header)
            excel_row = row + 2
            excel_col = col + 1
            
            # FILTRO: Solo actualizar columnas D (4), E (5) y F (6)
            if excel_col not in [4, 5, 6]:
                continue
            
            # Intentar obtener la celda
            cell = ws.cell(row=excel_row, column=excel_col)
            
            # Si es una celda combinada, buscar la celda principal
            if hasattr(cell, '__class__') and 'MergedCell' in str(cell.__class__):
                master_cell = _find_merge_master(ws, excel_row, excel_col)
                if master_cell:
                    cell = master_cell
                else:
                    logging.warning(f"Cannot find master cell for merged cell {cell_key}")
                    continue
            
            # Usar RichText
            try:
                logging.info(f"Procesando celda {cell_key}: {cell_data.value}")
                
                # Convertir HTML a RichText o texto plano
                cell_content = _parse_html_to_rich_text(cell_data.value)
                
                # Asignar contenido
                cell.value = cell_content
                
                # Configurar alineación y ajuste de texto
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                logging.info(f"✅ Celda {cell_key} actualizada exitosamente")
                
            except Exception as e:
                logging.warning(f"Cannot update cell {cell_key}: {str(e)}")
                # Fallback: asignar valor directo
                cell.value = cell_data.value
                continue
                
        except (ValueError, IndexError) as e:
            logging.warning(f"Error processing cell {cell_key}: {str(e)}")
            continue
             
def _update_cell_data_safe(ws, cell_data: Dict[str, models.CellData]):
    """Actualiza celdas con datos del frontend - CON RICH TEXT"""
    for cell_key, cell_data in cell_data.items():
        try:
            row, col = map(int, cell_key.split('-'))
            excel_row = row + 2
            excel_col = col + 1
            
            # FILTRO: Solo actualizar columnas D (4), E (5) y F (6)
            if excel_col not in [4, 5, 6]:
                continue
            
            # Intentar obtener la celda
            cell = ws.cell(row=excel_row, column=excel_col)
            
            # Si es una celda combinada, buscar la celda principal
            if hasattr(cell, '__class__') and 'MergedCell' in str(cell.__class__):
                master_cell = _find_merge_master(ws, excel_row, excel_col)
                if master_cell:
                    cell = master_cell
                else:
                    logging.warning(f"Cannot find master cell for merged cell {cell_key}")
                    continue
            
            # NUEVA LÓGICA: Usar RichText
            try:
                logging.info(f"Actualizando celda {cell_key}: {cell_data.value}")
                
                # Convertir HTML a RichText o texto plano
                cell_content = _parse_html_to_rich_text(cell_data.value)
                
                # Asignar contenido
                cell.value = cell_content
                
                # Configurar alineación y ajuste de texto
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                logging.info(f"✅ Celda {cell_key} actualizada exitosamente")
                
            except Exception as e:
                logging.warning(f"Cannot update cell {cell_key}: {str(e)}")
                # Fallback: asignar valor directo
                cell.value = cell_data.value
                continue
                
        except (ValueError, IndexError) as e:
            logging.warning(f"Error updating cell {cell_key}: {str(e)}")
            continue    
              
def _apply_cell_style(cell, styles: dict):
    """
    Aplica estilos extraídos del HTML a una celda de Excel
    """
    logging.info(f"=== APLICANDO ESTILOS ===")
    logging.info(f"Estilos a aplicar: {styles}")
    logging.info(f"Celda: {cell.coordinate}")
    
    try:
        # Obtener font actual o crear uno nuevo
        current_font = cell.font
        
        # Construir nuevo font con estilos combinados
        font_kwargs = {
            'name': current_font.name or 'Calibri',
            'size': current_font.size or 11,
            'bold': current_font.bold or False,
            'italic': current_font.italic or False,
            'underline': current_font.underline or 'none',
            'color': current_font.color
        }
        
        logging.info(f"Font actual: {font_kwargs}")
        
        # Aplicar estilos del HTML
        if 'bold' in styles:
            font_kwargs['bold'] = styles['bold']
            logging.info(f"✅ Aplicando NEGRITA: {styles['bold']}")
        
        if 'italic' in styles:
            font_kwargs['italic'] = styles['italic']
            logging.info(f"✅ Aplicando CURSIVA: {styles['italic']}")
        
        if 'color' in styles:
            try:
                color_hex = styles['color']
                if len(color_hex) == 6 and all(c in '0123456789ABCDEFabcdef' for c in color_hex):
                    font_kwargs['color'] = Color(rgb=color_hex)
                    logging.info(f"✅ Aplicando COLOR: {color_hex}")
                else:
                    logging.warning(f"❌ Color inválido: {color_hex}")
            except Exception as e:
                logging.error(f"❌ Error aplicando color {styles['color']}: {str(e)}")
        
        logging.info(f"Font final: {font_kwargs}")
        
        # Aplicar el font actualizado
        cell.font = Font(**font_kwargs)
        logging.info("✅ Font aplicado exitosamente")
        
        # Aplicar alineación y ajuste de texto
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
    except Exception as e:
        logging.error(f"❌ Error applying styles to cell: {str(e)}")
        import traceback
        logging.error(f"Traceback: {traceback.format_exc()}")

def _apply_styling(ws, template_config: models.TemplateConfig):
    """Aplica estilos generales al worksheet"""
    try:
        # Aplicar borde a todas las celdas con datos
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Determinar rango de datos
        max_row = template_config.tableConfig.get('numRows', 50) + 1
        max_col = template_config.tableConfig.get('numCols', 7)
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                
                # Altura de fila por defecto
                if row > 1:  # No aplicar al header
                    ws.row_dimensions[row].height = template_config.tableConfig.get('defaultRowHeight', 40) / 1.5
                    
    except Exception as e:
        logging.warning(f"Error applying general styling: {str(e)}")

def generate_filename(export_request: models.ExportExcelRequest) -> str:
    """Genera el nombre del archivo con formato: DD-MM-YYYY Cargue de Contenido [Proyecto] [Ciudad]"""
    from datetime import datetime
    from .secciones_service import _extract_city_slug

    info = export_request.template_info
    proyecto = (info.proyecto or "").lower()
    is_viajemos = "viajemos" in proyecto or "vjm" in proyecto
    site_name = "Viajemos" if is_viajemos else "Miles Car Rental"

    city_slug = _extract_city_slug(export_request.lp_url_slug)
    city_name = city_slug.capitalize()

    date_str = datetime.now().strftime("%d-%m-%Y")
    return f"{date_str} Cargue de Contenido {site_name} {city_name}.xlsx"